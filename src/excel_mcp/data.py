from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .exceptions import DataError
from .cell_utils import parse_cell_range
from .cell_validation import get_data_validation_for_cell

try:
    from xlcalculator import ModelCompiler, Evaluator
except ImportError:  # pragma: no cover - optional dependency
    ModelCompiler = None  # type: ignore[assignment]
    Evaluator = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

def _build_formula_evaluator(filepath: Path | str) -> Optional[Evaluator]:
    """Create an xlcalculator evaluator for the supplied workbook."""
    if ModelCompiler is None or Evaluator is None:
        logger.debug("xlcalculator not installed; formula evaluation disabled")
        return None
    resolved_path = Path(filepath).resolve(strict=False)
    try:
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(str(resolved_path))
        return Evaluator(model)
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning(
            "Failed to initialize xlcalculator evaluator for '%s': %s",
            filepath,
            exc,
        )
        return None

def _evaluate_formula_cell(
    evaluator: Optional[Evaluator],
    sheet_name: str,
    cell_address: str,
    row: int,
    column: int,
    fallback_loader: Optional[Callable[[int, int], Any]],
    default_value: Any,
) -> Any:
    """Evaluate a single formula cell and fall back to provided loader on failure."""
    computed_value: Any = None
    if evaluator is not None:
        normalized_sheet = sheet_name.replace("'", "''")
        tokens = [
            f"{sheet_name}!{cell_address}",
            f"'{normalized_sheet}'!{cell_address}",
        ]
        for token in tokens:
            try:
                computed_value = evaluator.evaluate(token)
                if computed_value is not None:
                    return computed_value
            except KeyError:
                # Try the next token variation
                continue
            except Exception as exc:  # pragma: no cover - library error handling
                logger.warning(
                    "Failed to evaluate formula in %s!%s via xlcalculator (%s): %s",
                    sheet_name,
                    cell_address,
                    token,
                    exc,
                )
                break
    if fallback_loader is not None:
        try:
            cached_value = fallback_loader(row, column)
            if cached_value is not None:
                return cached_value
        except Exception as exc:  # pragma: no cover - defensive fallback
            logger.debug("Formula fallback loader failed for %s!%s: %s", sheet_name, cell_address, exc)
    return default_value

def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> List[Dict[str, Any]]:
    """Read data from Excel range with optional preview mode"""
    try:
        wb = load_workbook(filepath, read_only=False)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries
                start_row, start_col = ws.min_row, ws.min_column
                end_row, end_col = ws.max_row, ws.max_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return []

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            if any(v is not None for v in row_data):
                data.append(row_data)

        wb.close()
        return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))

def write_data(
    filepath: str,
    sheet_name: Optional[str],
    data: Optional[List[List]],
    start_cell: str = "A1",
) -> Dict[str, str]:
    """Write data to Excel sheet with workbook handling
    
    Headers are handled intelligently based on context.
    """
    try:
        if not data:
            raise DataError("No data provided to write")
            
        wb = load_workbook(filepath)

        # If no sheet specified, use active sheet
        if not sheet_name:
            active_sheet = wb.active
            if active_sheet is None:
                raise DataError("No active sheet found in workbook")
            sheet_name = active_sheet.title
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validate start cell
        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if len(data) > 0:
            _write_data_to_worksheet(ws, data, start_cell)

        wb.save(filepath)
        wb.close()

        return {"message": f"Data written to {sheet_name}", "active_sheet": sheet_name}
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))

def _write_data_to_worksheet(
    worksheet: Worksheet, 
    data: List[List], 
    start_cell: str = "A1",
) -> None:
    """Write data to worksheet with intelligent header handling"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Write data
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                worksheet.cell(row=start_row + i, column=start_col + j, value=val)
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))

def read_excel_range_with_metadata(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    include_validation: bool = True,
    evaluate_formulas: bool = False
) -> Dict[str, Any]:
    """Read data from Excel range with cell metadata including validation rules.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell address
        end_cell: Ending cell address (optional)
        include_validation: Whether to include validation metadata
        evaluate_formulas: If True, evaluate formulas via xlcalculator (falls back to
                           cached values when available).
        
    Returns:
        Dictionary containing structured cell data with metadata
    """
    try:
        file_path = Path(filepath).resolve(strict=False)
        wb = load_workbook(file_path, read_only=False)
        formula_evaluator: Optional[Evaluator] = None
        wb_values = None
        ws_values = None
        fallback_loader: Optional[Callable[[int, int], Any]] = None
        fallback_loader_failed = False

        if evaluate_formulas:
            formula_evaluator = _build_formula_evaluator(file_path)

            def _get_cached_value(row: int, col: int) -> Any:
                nonlocal wb_values, ws_values, fallback_loader_failed
                if fallback_loader_failed:
                    return None
                if wb_values is None:
                    try:
                        wb_values = load_workbook(file_path, read_only=False, data_only=True)
                    except Exception as exc:  # pragma: no cover - IO issues
                        logger.warning(
                            "Unable to open workbook in data_only mode for fallback: %s",
                            exc,
                        )
                        fallback_loader_failed = True
                        return None
                    if sheet_name not in wb_values.sheetnames:
                        logger.warning(
                            "Sheet '%s' not found in data_only workbook during fallback",
                            sheet_name,
                        )
                        wb_values.close()
                        wb_values = None
                        fallback_loader_failed = True
                        return None
                    ws_values = wb_values[sheet_name]
                if ws_values is None:
                    return None
                return ws_values.cell(row=row, column=col).value

            fallback_loader = _get_cached_value

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries, but respect the provided start_cell
                end_row, end_col = ws.max_row, ws.max_column
                # If start_cell is 'A1' (default), we should find the true start
                if start_cell == 'A1':
                    start_row, start_col = ws.min_row, ws.min_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return {"range": f"{start_cell}:", "sheet_name": sheet_name, "cells": []}

        # Build structured cell data
        range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        range_data = {
            "range": range_str,
            "sheet_name": sheet_name,
            "cells": []
        }
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"

                value = cell.value
                formula_text: Optional[str] = None

                if cell.data_type == "f":
                    # openpyxl retains the formula string in cell.value for formula cells
                    if isinstance(cell.value, str):
                        formula_text = cell.value
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    # Defensive: some workbooks may expose formula as plain string even if data_type isn't 'f'
                    formula_text = cell.value

                if evaluate_formulas and formula_text is not None:
                    value = _evaluate_formula_cell(
                        evaluator=formula_evaluator,
                        sheet_name=sheet_name,
                        cell_address=cell_address,
                        row=row,
                        column=col,
                        fallback_loader=fallback_loader,
                        default_value=formula_text,
                    )
                elif evaluate_formulas and formula_evaluator is None and fallback_loader is not None:
                    fallback_value = fallback_loader(row, col)
                    if fallback_value is not None:
                        value = fallback_value

                cell_data = {
                    "address": cell_address,
                    "value": value,
                    "row": row,
                    "column": col
                }
                
                # Add validation metadata if requested
                if include_validation:
                    validation_info = get_data_validation_for_cell(ws, cell_address)
                    if validation_info:
                        cell_data["validation"] = validation_info
                    else:
                        cell_data["validation"] = {"has_validation": False}
                
                range_data["cells"].append(cell_data)

        wb.close()
        if wb_values is not None:
            wb_values.close()
        return range_data
        
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range with metadata: {e}")
        raise DataError(str(e))
